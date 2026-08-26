# -*- coding: utf-8 -*-
"""Construit le registre initial de la veille Hippolyte Bellangé (état des lieux du 25/08/2026)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DATE_VUE = "2026-08-25"

# (catégorie, titre, détails, source, type_source, pays, date_vente_ou_publi, estimation, prix, statut, url)
# Catégories : Peinture / Aquarelle-gouache / Dessin / Lithographie-estampe / Livre illustré / Publication / Objet-divers
# Statuts : À venir / En vente / Vendu / Invendu / Inconnu / Clôturé / Publié
ITEMS = [
    # ---------- MARCHÉ — EN VENTE ACTUELLEMENT (marchands, galeries, marketplaces) ----------
    ("Peinture", "Grenadier de la Garde en sentinelle", "HST signée « H. BELLANGE », datée 1829, 35 × 53 cm (67 × 85 encadrée)", "Galerie Dantan (Achicourt)", "Marchand", "FR", "", "", "2 200 €", "En vente", "https://en.dantan.store/product-page/joseph-louis-hippolyte-bellange-huile-sur-toile-grenadier-sign%C3%A9e-1800-1866"),
    ("Peinture", "The Four Militaries (attribué)", "HST 22 × 31 cm, milieu XIXe, attribution non documentée par une signature", "1stDibs — Gijsel Gallery (Genève)", "Marketplace", "CH", "", "", "1 152 $", "En vente", "https://www.1stdibs.com/art/paintings/portrait-paintings/hippolyte-bellange-four-militaries-hippolyte-bellange-antique-oil-on-canvas/id-a_12523192/"),
    ("Aquarelle-gouache", "Scène de rue", "Dessin/aquarelle signé « hte Bellangé », daté 1822, 22 × 17 cm, cadre doré XIXe", "Proantic — Galerie Mazarini (Lyon)", "Marchand", "FR", "", "", "2 350 €", "En vente", "https://www.proantic.com/en/1546464-bellange-hippolyte-1800-1866-quotstreet-scenequot-drawingwatercolor-signed-and-dated-be.html"),
    ("Aquarelle-gouache", "Le Camp du duc d'Orléans près de Compiègne", "Aquarelle signée « hte. Bgé. », datée 7bre 1836, 29 × 38 cm ; faite avec Raffet ; prov. probable vente après décès 1867", "Galerie La Nouvelle Athènes (Paris)", "Marchand", "FR", "2025-03-07", "", "Sur demande", "En vente", "https://lanouvelleathenes.fr/2025/03/07/hippolyte-bellange-1800-1866/"),
    ("Aquarelle-gouache", "La Vivandière de Wagram", "Aquarelle gouachée signée, 1862, 22 × 30,5 cm ; exposée Beaux-Arts 1867 n°157", "Galerie Christian Le Serbon (Paris)", "Marchand", "FR", "", "", "Sur demande", "En vente", "https://www.galerie-leserbon.fr/hippolyte-bellange-vivandiere-wagram/"),
    ("Dessin", "Bataille de Sébastopol", "Lavis brun signé, annoté au dos « Sébastopol 1855 », 25,5 × 82,5 cm, pli central", "Proantic — Galerie Mazarini (Lyon)", "Marchand", "FR", "", "", "1 450 €", "En vente", "https://www.proantic.com/en/657821-bellange-hippolyte-1800-1866-quotbattle-of-sebastopolquot-drawing-brown-wash-signed-anno.html"),
    ("Dessin", "Figure", "Encre de Chine, 7 × 5,5 cm, milieu XIXe", "1stDibs — Wallector (Rome)", "Marketplace", "IT", "", "", "Sur demande", "En vente", "https://www.1stdibs.com/art/drawings-watercolor-paintings/figurative-drawings-watercolors/hippolyte-bellange-figure-original-drawing-china-ink-hippolyte-bellange-mid-19th-century/id-a_10777412/"),
    ("Lithographie-estampe", "Album « Souvenirs militaires » (1834), Gihaut — 12 planches + couverture", "En feuilles, 340 × 510 mm : Jemmapes, Mont Saint-Bernard, Marengo, Arcole, Moskowa, Montereau, Bastan, Landrecies, Retour de l'île d'Elbe, Guadarrama, Valmy, camp de Boulogne", "Estampes MAS", "Marchand", "FR", "", "", "2 000 €", "En vente", "https://www.estampes-mas.fr/produit/bellange-souvenirs-militaires-album-1834/"),
    ("Lithographie-estampe", "Bataille de la Moskowa (modèle de coloris)", "Litho vers 1832, coupée au sujet, 185 × 300 mm", "Estampes MAS", "Marchand", "FR", "", "", "300 €", "En vente", "https://www.estampes-mas.fr/produit/bellange-bataille-de-la-moskowa/"),
    ("Lithographie-estampe", "Bataille de Marengo (modèle de coloris)", "Litho signée/datée 1832, 190 × 310 mm", "Estampes MAS", "Marchand", "FR", "", "", "Non affiché", "En vente", "https://www.estampes-mas.fr/produit/bellange-bataille-de-marengo/"),
    ("Lithographie-estampe", "La Garde meurt mais ne se rend pas", "Litho 1849, 505 × 370 mm, mouillure angle (Aveline 409)", "Estampes MAS", "Marchand", "FR", "", "", "Non affiché", "En vente", "https://www.estampes-mas.fr/produit/bellange-la-garde-meurt-mais-ne-se-rend-pas/"),
    ("Lithographie-estampe", "La Vivandière de Béranger", "Litho Villain/Gihaut, 245 × 175 mm", "Estampes MAS", "Marchand", "FR", "", "", "30 €", "En vente", "https://www.estampes-mas.fr/produit/bellange-la-vivandiere-de-beranger/"),
    ("Lithographie-estampe", "Dédié à Madame Vve Charlet (1846)", "Litho Villain/Gihaut, 42 × 57 cm, légère humidité", "Pamono — vendeur Italie", "Marketplace", "IT", "", "", "397 $", "En vente", "https://www.pamono.com/hippolyte-bellange-dedie-a-madame-vve-charlet-original-lithograph-1846"),
    ("Lithographie-estampe", "Tiens bon, Turc ! Nous voilà, mon brave (Crimée, v. 1854)", "1re édition, Auguste Bry / Delarue, feuille 45 × 58,5 cm", "AbeBooks — Wittenborn Art Books (San Francisco)", "Marketplace", "US", "", "", "264,82 €", "En vente", "https://www.abebooks.fr/art-affiches/bon-Turc-brave-Guerre-Crim%C3%A9e-1853-1856/30932387763/bd"),
    ("Lithographie-estampe", "Tenez voyez-vous, Mr le Curé… (Albums lithographiques n°8, 1834)", "Litho coloriée d'époque, image 17 × 20,5 cm", "AbeBooks — Antiquariat Friederichsen (Hambourg)", "Marketplace", "DE", "", "", "300 €", "En vente", "https://www.abebooks.fr/art-affiches/Tenez-Voyez-Vous-Cure-v%60la.l%60Pere-Eternel-Altkolorierte/15341477405/bd"),
    ("Lithographie-estampe", "Le retour de l'Armée (Villain, v. 1835)", "Litho coloriée", "AbeBooks — Antiquariat Friederichsen (Hambourg)", "Marketplace", "DE", "", "", "Non relevé", "En vente", "https://www.abebooks.fr/art-affiches/retour-Armee-Altkolorierte-Lithographie-Villain-Bellange/15341477401/bd"),
    ("Lithographie-estampe", "Ces messieurs prennent leur café, à c'qui paraît (Gihaut, années 1830)", "1re édition, 32 × 36,5 cm (Adeline 377)", "AbeBooks — Wittenborn Art Books", "Marketplace", "US", "", "", "132,41 €", "En vente", "https://www.abebooks.fr/art-affiches/messieurs-prennent-caf%C3%A9-cqui-parait-First/30932390553/bd"),
    ("Lithographie-estampe", "Napoléon à Montereau (1828)", "Litho originale 24,5 × 34 cm", "1stDibs — Wallector (Rome)", "Marketplace", "IT", "", "", "Sur demande", "En vente", "https://www.1stdibs.com/art/prints-works-on-paper/figurative-prints-works-on-paper/hippolyte-bellange-napoleon-montereau-original-lithograph-h-bellange-1828/id-a_4031282/"),
    ("Lithographie-estampe", "Vive le Vin (Engelmann, v. 1825)", "Litho à la teinte, feuille 34,8 × 26,7 cm (Béraldi p. 17)", "1stDibs — Thomas French Fine Art (Ohio)", "Marketplace", "US", "", "", "Sur demande", "En vente", "https://www.1stdibs.com/art/prints-works-on-paper/hippolyte-bellange-vive-le-vin/id-a_15017462/"),
    ("Lithographie-estampe", "Grenadier (Villain, 1824), pl. 3", "Litho coloriée 22,7 × 29,6 cm", "AbeBooks — ThePrintsCollector (NL)", "Marketplace", "NL", "", "", "148,43 $", "En vente", "https://www.abebooks.com/art-prints/Antique-Print-GRENADIER-MILITARY-INFANTRY-BAYONET-PL-3-Bellange-Villain-1824/22589589902/bd"),
    ("Lithographie-estampe", "Bonaparte à Toulon (eau-forte, 1837, Norvins)", "22 × 14 cm", "Pamono — vendeur Italie", "Marketplace", "IT", "", "", "230 $", "En vente", "https://www.pamono.com/hippolyte-bellange-napoleon-bonaparte-in-toulon-1837-etching"),
    ("Lithographie-estampe", "Le Mendiant (Alliance des Arts)", "Litho", "eBay.fr — lefennec205 (Pluvigner)", "Marketplace", "FR", "", "", "25 €", "En vente", "https://www.ebay.fr/itm/264629880742"),
    ("Livre illustré", "Adeline, « Hippolyte Bellangé et son œuvre » (Quantin 1880) — 1/50 sur Hollande", "Maroquin rouge signé Pagnant, enrichi de 12 dessins originaux d'Adeline d'après Bellangé", "Livre-rare-book — ILLIBRAIRIE (Genève)", "Marchand", "CH", "", "", "2 138,35 €", "En vente", "https://www.livre-rare-book.com/book/20676660/4280"),
    ("Livre illustré", "Norvins, « Histoire de Napoléon », 21e éd. (Furne 1868), ill. Raffet, Charlet, Bellangé, Yan' Dargent", "In-4, demi-veau aubergine, rousseurs", "AbeBooks — Mesnard Comptoir du Livre Ancien", "Marketplace", "FR", "", "", "108,16 $", "En vente", "https://www.abebooks.com/Histoire-Napol%C3%A9on-Norvins-Vingt-uni%C3%A8me-%C3%A9dition/19905379626/bd"),
    ("Livre illustré", "Collection des types de tous les corps et des uniformes militaires… (1844), 50 pl. coloriées", "« A superb copy » — libraire et prix non lisibles", "viaLibri", "Agrégateur", "", "", "", "Non lisible", "En vente", "https://www.vialibri.net/years/books/361033254/1844-bellange-m-hippolyte-collection-des-types-de-tout-les"),

    # ---------- MARCHÉ — VENDU / CLÔTURÉ RÉCEMMENT (marchands) ----------
    ("Aquarelle-gouache", "Le départ du conscrit", "Aquarelle signée/datée 1846, 25,7 × 30,5 cm encadrée", "Nahum Gallery (Vitry-sur-Seine)", "Marchand", "FR", "", "", "450 €", "Vendu", "https://nahumgallery.com/en-inter/products/hippolyte-bellange-aquarelle-le-depart-du-conscrit"),
    ("Aquarelle-gouache", "Soldat blessé", "Aquarelle 25,5 × 32 cm", "Nahum Gallery", "Marchand", "FR", "", "", "250 €", "Vendu", "https://nahumgallery.com/en-inter/products/soldat-bellange"),
    ("Dessin", "Un ruffian à cheval", "Crayon rehaussé d'aquarelle et encre, cadre XXe", "Proantic — Antiquités de la Sapinière", "Marchand", "FR", "", "", "550 €", "Vendu", "https://www.proantic.com/en/1157734-dessin-d039-hippolyte-bellange-quot1800-1866-un-ruffian-a-cheval.html"),
    ("Dessin", "Triple Study of Soldier with Rifle", "Dessin ; annonce clôturée par le vendeur le 13 juin 2026", "eBay.com — museumcoin (Baltimore)", "Marketplace", "US", "2026-06-13", "", "295,95 $", "Clôturé", "https://www.ebay.com/itm/235442304483"),
    ("Lithographie-estampe", "Passage du pont d'Arcole (modèle de coloris)", "Litho 1832", "Estampes MAS", "Marchand", "FR", "", "", "", "Vendu", "https://www.estampes-mas.fr/produit/bellange-passage-du-pont-darcole/"),

    # ---------- ENCHÈRES 2025-2026 ----------
    ("Peinture", "Zouaves de la Garde impériale lors d'une halte, après la campagne de Crimée", "HST signée b.g., 46 × 54 cm, restaurations, cadre Binant ; prov. Dr Nacquart ; peut-être n°447 du catalogue (anc. coll. Baroche)", "Osenat — L'Empire à Fontainebleau, lot 234", "Maison de vente", "FR", "2025-12-07", "3 000 – 4 000 €", "Non publié", "Invendu probable", "https://www.osenat.com/en/lot/169041/31518386-hippolyte-bellange-1800-1866-zouaves-of-the-imperial-guard"),
    ("Peinture", "The Barber of the Pays de Caux (d'après H. Bellangé)", "Huile sur bois 33 × 41 cm ; passé le 28/10/2025 puis représenté le 05/05/2026 (maison masquée, paywall Artnet)", "Artnet (maison non affichée)", "Agrégateur", "", "2026-05-05", "Paywall", "Paywall", "Inconnu", "https://www.artnet.com/artists/hippolyte-bellangé/the-barber-of-the-pays-de-caux-IcD1M9GYh0m_zzhxIMtxTQ2"),
    ("Peinture", "Convoi militaire français arrivant dans un village (1829)", "Technique non lue (probable HST) 32 × 46 cm", "Carlo Bonte Auctions (Bruges) via Invaluable", "Maison de vente", "BE", "2025-10-14", "2 000 – 3 000 €", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/hippolyte-bellange-1800-1866-french-military-conv-316-c-8a04df38e7"),
    ("Peinture", "Scène de bataille, HST signée", "66 × 97 cm, lot 560", "DVC (Gand) via the-saleroom / Invaluable", "Maison de vente", "BE", "2025-06-22", "700 – 1 200 €", "Non affiché", "Inconnu", "https://www.the-saleroom.com/en-gb/auction-catalogues/dvc/catalogue-id-dvc-nv10062/lot-f6dd7cfc-578a-4c60-b228-b2f400b3e13b"),
    ("Peinture", "Deux tableaux encadrés, hommes en uniforme (lot 110)", "Description issue de la vignette, technique non précisée — page de lot en 410", "Bailleul & Nentas (Bayeux) via Interencheres", "Maison de vente", "FR", "2026-07-14", "200 – 300 €", "Non affiché", "Inconnu", "https://www.interencheres.com/art-decoration/grande-vente-du-14-juillet-a-14h-684108"),
    ("Dessin", "Mousquetaire (Musketeer)", "Dessin, vente 83 lot 6352", "Bubb Kuyper (Haarlem) via Invaluable", "Maison de vente", "NL", "2025-11-28", "80 – 100 €", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/bellange-hippolyte-1800-1866-musketeer-6352-c-d14d484cda"),
    ("Dessin", "Overgrown monument", "Plume ; représenté à répétition (11/2023, 03/2024, 04/2024, 06/2025)", "Fichter Kunsthandel (Francfort) via Invaluable", "Maison de vente", "DE", "2025-06-14", "1 200 – 1 600 €", "Non affiché", "Invendu probable", "https://www.invaluable.com/auction-lot/h-bellange-1800-1866-overgrown-monument-around-18-15-c-67343b8acf"),
    ("Dessin", "Caricature — Cavalryman at the fire", "Représenté depuis 02/2023", "Fichter Kunsthandel via Invaluable", "Maison de vente", "DE", "2025-05-18", "60 – 100 €", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/h-bellange-1800-1866-caricature-cavalryman-at-the-620-c-4664ab98f3"),
    ("Dessin", "Caricature — Woman and soldier", "", "Fichter Kunsthandel via Invaluable", "Maison de vente", "DE", "2025-05-18", "60 – 100 €", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/h-bellange-1800-1866-caricature-woman-and-soldier-619-c-3d84434bb8"),
    ("Dessin", "Caricature — Couple fishing", "", "Fichter Kunsthandel via Invaluable", "Maison de vente", "DE", "2025-05-18", "60 – 100 €", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/h-bellange-1800-1866-caricature-couple-fishing-li-618-c-9bc4d26ac8"),
    ("Lithographie-estampe", "Dragon et Sapeur (Types de tous les corps, 1844)", "16,5 × 12 cm, signée sur la pierre", "Sopocki Dom Aukcyjny via OneBid", "Maison de vente", "PL", "2025-04-05", "215 €", "258 €", "Vendu", "https://onebid.fr/fr/arts-graphiques-et-dessins-hippolyte-bellange-1800-1866-dragon-et-sapeur/2724782"),
    ("Lithographie-estampe", "Général napoléonien", "", "Sopocki Dom Aukcyjny via OneBid", "Maison de vente", "PL", "2025-04-05", "", "258 €", "Vendu", "https://onebid.fr/fr/arts-graphiques-et-dessins-hippolyte-bellange-1800-1866-general-napoleonien/2724783"),
    ("Lithographie-estampe", "Portrait à cheval du prince Eugène de Beauharnais", "", "Sopocki Dom Aukcyjny via OneBid", "Maison de vente", "PL", "2025-04-05", "", "258 €", "Vendu", "https://onebid.fr/fr/arts-graphiques-et-dessins-hippolyte-bellange-1800-1866-portrait-a-cheval-du-prince-eugen-beauharnais/2724784"),
    ("Lithographie-estampe", "Soldats de la Garde de Napoléon", "", "Sopocki Dom Aukcyjny via OneBid", "Maison de vente", "PL", "2025-04-05", "", "258 €", "Vendu", "https://onebid.fr/fr/arts-graphiques-et-dessins-hippolyte-bellange-1800-1866-soldats-de-la-garde-de-napoleon/2724785"),
    ("Lithographie-estampe", "Chasseur à cheval 1812 / Officier d'ordonnance (Château de Lesko, 2 lots)", "Lithos coloriées 19 × 15 cm, série Dubochet 1844", "Mercari Polonia (Varsovie) via OneBid", "Maison de vente", "PL", "2026-03-14", "25 € chacun", "Sans enchère", "Invendu", "https://onebid.fr/fr/arts-graphiques-et-dessins-lesko-castle-hippolyte-bellange-lithographie-du-19eme-siecle-horse-chaser-1812-chasseur-a-cheval-1812-famille-krasicki/3299367"),
    ("Lithographie-estampe", "4 estampes Bellangé", "Lots 2634245-2634248", "Dom Aukcyjny Ostoya (Varsovie) via OneBid", "Maison de vente", "PL", "2025-02-08", "47 € chacun", "Sans enchère", "Invendu", "https://onebid.fr/fr/artist/aukcja/Hippolyte-Bellange"),
    ("Lithographie-estampe", "Two lithographs (lot 45)", "Old Master, British & European Pictures", "Roseberys (Londres) via Invaluable", "Maison de vente", "UK", "2025-07-09", "100 – 150 £", "Non affiché", "Inconnu", "https://www.invaluable.com/auction-lot/joseph-louis-hippolyte-bellange-french-1800-1866--45-c-f8847c68b8"),
    ("Lithographie-estampe", "Napoléon, deux gravures d'après Vernet et Bellangé (lot 24)", "", "Hôtel des Ventes Giraudeau (Tours)", "Maison de vente", "FR", "2025-02-17", "", "65 € (79,82 € TTC)", "Vendu", "https://docs.prod-indb.io/2025/02/18/223435_613271382_99c0c394324e4f7b511d2a50999ef658.pdf"),
    ("Livre illustré", "Janin, « La Normandie » (Bourdin 1844), hors-textes d'après Bellangé et al. (lot 242)", "", "Maison non identifiée (PDF prod-indb)", "Maison de vente", "FR", "2025-09-11", "", "20 €", "Vendu", "https://docs.prod-indb.io/2025/09/24/082234_574046984_49287f43269ccc9398973d69ccaf6b8f.pdf"),
    ("Livre illustré", "Militaria : livres et albums (12 pièces) dont « Costumes militaires français » de Bellangé (lot 1099)", "Gallery Closure Hassfurther II", "im Kinsky (Vienne) via the-saleroom", "Maison de vente", "AT", "2025-05-21", "Non affiché", "Non affiché", "Inconnu", "https://www.the-saleroom.com/en-gb/auction-catalogues/imkinsky/catalogue-id-auktio1-10036/lot-292d5c95-4d81-4f10-8848-b2cf0180ee76"),
    ("Objet-divers", "[DOCUMENTATION] lot 400", "Fiche artiste Millon, description non affichée", "Millon", "Maison de vente", "FR", "2025-10-30", "", "230 €", "Vendu", "https://www.millon.com/createurs/hippolyte-bellange"),
    ("Objet-divers", "Lot « à venir » signalé par MutualArt (1 upcoming)", "Détail verrouillé — à identifier", "MutualArt", "Agrégateur", "", "", "", "", "À venir", "https://www.mutualart.com/Artist/Hippolyte-Bellange/CB676CE910C75B49/artworks-for-sale?Type=Upcoming"),

    # ---------- RÉFÉRENCES DE COTE (résultats antérieurs marquants) ----------
    ("Peinture", "L'Empereur Napoléon Ier à Wagram, juillet 1809", "HST 57 × 72,5 cm, signée/datée 1841, Salon 1842 n°75 (réf. cote)", "Osenat — L'Empire à Fontainebleau, lot 212", "Maison de vente", "FR", "2015-04-12", "10 000 – 12 000 €", "12 500 €", "Vendu", "https://www.osenat.com/lot/21937/4938999-joseph-louis-hippolyte-bellange-paris-1800-1866-lempereur"),
    ("Peinture", "The Surrender of Aboukir Fort", "Huile sur panneau 162 × 212 cm (réf. cote)", "Artcurial via MutualArt", "Maison de vente", "FR", "2015-11-13", "", "57 763 $", "Vendu", "https://www.mutualart.com/Artwork/THE-SURRENDER-OF-ABOUKIR-FORT/439583F983A5E893"),
    ("Peinture", "Scène de bataille panoramique (Moskowa)", "HST 65 × 98 cm, signée (réf. cote)", "Aguttes — vente Empire bicentenaire, lot 30", "Maison de vente", "FR", "2021-06", "", "6 500 € TTC", "Vendu", "https://www.aguttes.com/en/lot/112344/15053759-hippolyte-bellange-paris-18001"),
    ("Peinture", "Grenadiers à pied de la Garde en faction près d'un moulin, paysage de neige", "Huile sur panneau 1854, 31 × 24 cm, coll. Bernard Franck (réf. cote)", "Thierry de Maigret — Souvenirs historiques, lot 131", "Maison de vente", "FR", "", "800 – 1 500 €", "2 300 €", "Vendu", "https://www.thierrydemaigret.com/en/lot/22008/4766805"),
    ("Peinture", "Convoi militaire", "HST (réf. cote)", "Vanderkindere via Invaluable", "Maison de vente", "BE", "2024-11-13", "2 000 – 2 500 €", "Non lu", "Inconnu", "https://www.invaluable.com/auction-lot/bellange-hippolyte-1800-1866-tableaux-huile-sur-t-388-c-cff495fbf5"),
    ("Peinture", "Réquisition de vivres au pied d'un moulin (1851)", "HST 32,5 × 41 cm (réf. cote)", "Briscadieu (Bordeaux)", "Maison de vente", "FR", "2021-06-05", "", "1 100 €", "Vendu", "https://www.briscadieu-bordeaux.com//en/lot/113985/15097257"),
    ("Dessin", "La Retraite de Constantine (1845), annoté par le général Changarnier", "Dessin, coll. napoléonienne du Palais de Monaco, lot 39 (réf. cote)", "Osenat / Giquello", "Maison de vente", "FR", "2024-11", "800 – 1 200 €", "4 000 €", "Vendu", "https://www.giquelloetassocies.fr/lot/21004/4364429-hyppolite-bellange-1800-1866-la-retraite-de-constantine"),
    ("Dessin", "Le Soldat de l'An II (1853)", "Dessin aquarellé, coll. Palais de Monaco, lot 29 (réf. cote)", "Osenat", "Maison de vente", "FR", "2024-11", "300 – 500 €", "700 €", "Vendu", "https://www.osenat.com/lot/21004/4364419-hyppolite-bellange-18001866le"),
    ("Aquarelle-gouache", "Grenadier à pied de la Garde impériale (1859)", "Dessin aquarellé 40 × 29 cm (réf. cote)", "Osenat", "Maison de vente", "FR", "2024-11-19", "", "688 € TTC", "Vendu", "https://www.osenat.com//en/lot/78093/6514370"),
    ("Livre illustré", "Uniformes de l'armée française depuis 1815 (108 lithos coloriées) + 5 lithos", "(réf. cote)", "Ader, lot 69", "Maison de vente", "FR", "2023-05-30", "", "750 €", "Vendu", "https://www.ader-paris.fr/lot/136271/21652865-bellange-hippolyte-uniformes-de-larmee-francaise-depuis-1815"),

    # ---------- PUBLICATIONS & ACTUALITÉS 2024-2026 ----------
    ("Publication", "Solène Sazio, « Hippolyte Bellangé (1800-1866), imagier de la légende napoléonienne »", "Article, Napoleonica. La Revue 2024/3 n°50, p. 159 sq. — issu de la thèse (Rouen 2018)", "Cairn — Napoleonica", "Publication", "FR", "2024", "", "", "Publié", "https://shs.cairn.info/revue-napoleonica-la-revue-2024-3-page-159?lang=fr"),
    ("Publication", "Dépôt HAL de l'article Sazio (hal-04995757)", "Version auteur en libre accès", "HAL", "Publication", "FR", "2025", "", "", "Publié", "https://hal.science/hal-04995757v1"),
    ("Publication", "CV HAL de Solène Sazio — auteure de référence", "13 documents ; communication 2026 « Dessiner l'expédition du Mexique (1861-1867) »", "HAL", "Publication", "FR", "2026", "", "", "Publié", "https://cv.hal.science/solene-sazio"),
    ("Publication", "Claudia Bonnafoux, « Le mal du pays », Album patriotique (1833) pl. 10", "Notice iconographique, Fondation Napoléon, février 2024", "napoleon.org", "Publication", "FR", "2024-02", "", "", "Publié", "https://www.napoleon.org/histoire-des-2-empires/iconographie/le-mal-du-pays-album-lithographique-par-hippolyte-bellange/"),
    ("Publication", "Notice d'autorité Musée d'Orsay n°37312 (mise à jour 7 avril 2026)", "Bellangé daguerréotypiste amateur ; renvoi au portrait PHO 1995 6 246", "Musée d'Orsay", "Publication", "FR", "2026-04-07", "", "", "Publié", "https://www.musee-orsay.fr/en/ressources/artists-personalities-catalog/hippolyte-bellange-37312"),
    ("Publication", "Notice Louvre RF 60 — Un jour de revue sous l'Empire (1810), 1862", "Historique complet ; prêté à l'expo « Victoire ! » (Musée de l'Armée 2023-24)", "Louvre collections", "Publication", "FR", "", "", "", "Publié", "https://collections.louvre.fr/en/ark:/53355/cl010065827"),
    ("Publication", "Exposition « Victoire ! La fabrique des héros » (catalogue Leluc et al.)", "Musée de l'Armée, 11 oct. 2023 – 28 janv. 2024 — prêt du Louvre RF 60", "Musée de l'Armée", "Publication", "FR", "2023-10-11", "", "", "Publié", "https://www.musee-armee.fr/au-programme/expositions/expositions-passees.html"),
    ("Publication", "La Tribune de l'Art — « Hippolyte Bellangé (1800-1866) » (id 4582)", "Contenu et date non vérifiés (site inaccessible aux robots) — à ouvrir manuellement", "La Tribune de l'Art", "Publication", "FR", "", "", "", "À vérifier", "https://www.latribunedelart.com/hippolyte-bellange-1800-1866"),
    ("Publication", "Solène Sazio, thèse « Hippolyte Bellangé (1800-1866), reconnaissance et oubli d'un artiste aux origines de la légende napoléonienne »", "Université de Rouen Normandie, soutenue 23 mars 2018, dir. Marec / Le Men (socle)", "theses.fr", "Publication", "FR", "2018-03-23", "", "", "Publié", "https://theses.fr/2018NORMR021"),
    ("Publication", "Francis Wey, « Exposition des œuvres d'Hippolyte Bellangé… étude biographique » (1867)", "Première monographie (socle), Gallica", "Gallica", "Publication", "FR", "1867", "", "", "Publié", "https://gallica.bnf.fr/ark:/12148/bpt6k6555744x"),
]

wb = Workbook()

# ---------- Feuille Registre ----------
ws = wb.active
ws.title = "Registre"
headers = ["ID", "Date vue", "Catégorie", "Titre", "Détails", "Source", "Type source", "Pays",
           "Date vente / publication", "Estimation", "Prix", "Statut", "URL", "Intérêt (à remplir)", "Mes notes"]
ws.append(headers)
for i, it in enumerate(ITEMS, start=1):
    cat, titre, det, src, typ, pays, dv, est, prix, statut, url = it
    ws.append([f"B{i:04d}", DATE_VUE, cat, titre, det, src, typ, pays, dv, est, prix, statut, url, "", ""])

widths = [8, 11, 20, 52, 55, 36, 16, 6, 14, 16, 16, 15, 60, 16, 30]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="3B3B6D")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    row[12].hyperlink = row[12].value
    row[12].font = Font(name="Arial", size=10, color="0563C1", underline="single")
    for idx in (13, 14):
        row[idx].fill = PatternFill("solid", fgColor="FFF9C4")
ws.freeze_panes = "D2"
tab = Table(displayName="Registre", ref=f"A1:{get_column_letter(len(headers))}{len(ITEMS)+1}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
ws.add_table(tab)

# ---------- Feuille Sources ----------
ws2 = wb.create_sheet("Sources")
ws2.append(["Famille", "Source", "URL de surveillance", "Accès robot", "Alerte e-mail à créer", "Notes"])
SOURCES = [
    ("Maisons de vente FR", "Osenat (Fontainebleau) — ventes Empire", "https://www.osenat.com", "OK (fiches de lots)", "Oui — newsletter/alerte « Bellangé »", "Source n°1 pour les peintures ; catalogue paginé partiellement lisible"),
    ("Maisons de vente FR", "Millon — fiche artiste", "https://www.millon.com/createurs/hippolyte-bellange", "OK", "Oui", "Résultats sans URL de lot"),
    ("Maisons de vente FR", "Ader, Aguttes, Tajan, Thierry de Maigret, Coutau-Bégarie, Rossini, De Baecque, Libert, Émeraude, Giquello, Rouillac…", "recherche « Bellangé » sur chaque site", "OK", "Oui quand disponible", "Fiches lisibles"),
    ("Plateformes FR", "Interencheres", "https://www.interencheres.com/recherche?q=Bellangé", "BLOQUÉ (404/410)", "OUI — alerte mot-clé « Bellangé » (indispensable)", "Couvre les maisons régionales (Bayeux, Évreux, Tours, Joué…)"),
    ("Plateformes FR", "Drouot.com", "https://drouot.com/fr/recherche?q=Bellangé", "BLOQUÉ", "OUI — alerte « Bellangé »", ""),
    ("Plateformes FR", "Gazette Drouot", "https://www.gazette-drouot.com/recherche?q=Bellangé", "BLOQUÉ", "Newsletter", "Ventes à venir + résultats commentés"),
    ("Agrégateurs", "Invaluable — page artiste", "https://www.invaluable.com/artist/bellange-hippolyte-5d8abb2azf/sold-at-auction-prices/", "OK (liste)", "OUI — alerte mot-clé", "La plus complète à l'international ; pages de lot vides (JS)"),
    ("Agrégateurs", "OneBid — page artiste", "https://onebid.fr/fr/artist/aukcja/Hippolyte-Bellange", "OK (prix + statut)", "Oui", "Marché polonais des lithos"),
    ("Agrégateurs", "Artnet — résultats", "https://www.artnet.com/artists/hippolyte-bellang%C3%A9/auction-results", "OK (index) / prix paywall", "Non", "Utile pour repérer les passages (ex. Barbier du pays de Caux)"),
    ("Agrégateurs", "the-saleroom — price guide", "https://www.the-saleroom.com/en-gb/price-guide/hippolyte-bellangé", "OK (liste)", "Oui (alerte lot)", ""),
    ("Agrégateurs", "MutualArt", "https://www.mutualart.com/Artist/Hippolyte-Bellange/CB676CE910C75B49", "Compteurs seulement", "Oui (gratuit, « follow artist »)", "Signale les lots à venir"),
    ("Agrégateurs", "LiveAuctioneers — price guide", "https://www.liveauctioneers.com/price-guide/hippolyte-bellange/10477/", "OK", "Oui (keyword alert)", "Fichter Kunsthandel (DE) y vend régulièrement"),
    ("Agrégateurs", "LotSearch, Barnebys, Artprice", "", "BLOQUÉ / paywall", "Barnebys : alerte gratuite", ""),
    ("Marketplaces", "eBay.fr / eBay.com", "recherche « Bellangé » dans Art > Estampes / Dessins", "BLOQUÉ (recherche) ; fiches /itm/ lisibles", "OUI — recherche sauvegardée « Bellangé » avec alerte", "Beaucoup d'annonces expirées indexées"),
    ("Marketplaces", "Catawiki", "https://www.catawiki.com/fr/s?q=Bellangé", "BLOQUÉ", "OUI — alerte « Bellangé »", ""),
    ("Marketplaces", "Delcampe", "https://www.delcampe.net/fr/collections/search?term=Bellangé", "BLOQUÉ", "OUI — alerte « Bellangé »", "Cartes/lithos/livres"),
    ("Marketplaces", "Leboncoin", "https://www.leboncoin.fr/recherche?text=bellangé", "BLOQUÉ", "OUI — alerte « bellangé »", "Attention homonymes (ébénistes Bellangé)"),
    ("Marketplaces", "1stDibs, Pamono, Selency", "recherche « Bellangé »", "OK (fiches ; prix 1stDibs à la connexion)", "1stDibs : « follow » possible", ""),
    ("Marchands", "Proantic", "https://www.proantic.com/recherche?q=Bellangé", "Recherche bloquée ; fiches lisibles", "OUI — alerte « Bellangé »", "Galerie Mazarini y est active"),
    ("Marchands", "Anticstore, Les Atamanes, Galerie Napoléon, Estampes MAS, Galerie Dantan, Le Serbon, La Nouvelle Athènes, Nahum", "sites propres", "Variable (Galerie Napoléon 403)", "Newsletters", "Galerie Napoléon = probablement le plus gros stock d'estampes, à consulter à la main"),
    ("Livres", "Livre-rare-book, AbeBooks, viaLibri, Galaxidion", "recherche auteur « Bellangé »", "Fiches lisibles ; recherche LRB bloquée", "AbeBooks : alerte « Bellangé » ; viaLibri : « want »", "Norvins, Uniformes 1844, Adeline 1880, Wey 1867"),
    ("Publications", "Cairn / Napoleonica", "https://shs.cairn.info/resultats_recherche.php?searchTerm=Bellangé", "BLOQUÉ", "Non", "À ouvrir à la main ou via Gmail (Google Scholar)"),
    ("Publications", "HAL / theses.fr / DUMAS / Persée / OpenEdition", "voir URL dans la skill", "HAL bloqué ; autres OK", "Google Scholar : alerte « Hippolyte Bellangé »", "Suivre Solène Sazio"),
    ("Publications", "Google Alerts", "https://www.google.com/alerts", "—", "OUI — « Hippolyte Bellangé » (hebdo)", "Filtrer homonymes : Jacques Bellange, ébénistes Bellangé, Eugène Bellangé"),
    ("Publications", "POP/Joconde, Gallica, Louvre, Orsay, Wikidata (historique)", "voir URL dans la skill", "OK", "Non", "Nouvelles notices, acquisitions, prêts"),
    ("Publications", "Musées : MBA Rouen, Musée de l'Armée, Malmaison, Versailles, Fondation Napoléon", "pages expositions", "OK", "Newsletters", ""),
    ("Presse", "La Tribune de l'Art, Journal des Arts, Connaissance des Arts, Souvenir Napoléonien", "recherche « Bellangé »", "Tribune de l'Art bloqué", "Newsletters", ""),
]
for s in SOURCES:
    ws2.append(list(s))
for c, w in enumerate([18, 50, 60, 24, 34, 50], start=1):
    ws2.column_dimensions[get_column_letter(c)].width = w
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws2.freeze_panes = "A2"

# ---------- Feuille Journal ----------
ws3 = wb.create_sheet("Journal")
ws3.append(["Date du passage", "Type", "Entrées ajoutées", "Nouveautés signalées", "Brief envoyé", "Remarques"])
ws3.append([DATE_VUE, "État des lieux initial", len(ITEMS), "—", "Brief n°0", "Balayage complet (4 explorateurs, ~500 requêtes). Sources bloquées : Interencheres, Drouot, eBay, Leboncoin, Catawiki, Delcampe, Cairn, HAL, Gazette Drouot, Tribune de l'Art."])
for c, w in enumerate([16, 24, 16, 20, 14, 80], start=1):
    ws3.column_dimensions[get_column_letter(c)].width = w
for cell in ws3[1]:
    cell.font = header_font
    cell.fill = header_fill
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# ---------- Feuille Légende ----------
ws4 = wb.create_sheet("Légende")
LEG = [
    ["Registre de veille — Hippolyte Bellangé (1800-1866)"],
    [""],
    ["Fonctionnement", "Chaque passage hebdomadaire ajoute les nouvelles entrées (une ligne par lot / annonce / publication) et ne signale dans le brief que les URL absentes du registre."],
    ["Colonnes à remplir par Jérôme", "« Intérêt » (fond jaune) : Oui / Non / Acheté / Suivi. « Mes notes » : libre. Tout le reste est rempli automatiquement."],
    ["Catégories", "Peinture · Aquarelle-gouache · Dessin · Lithographie-estampe · Livre illustré · Publication · Objet-divers"],
    ["Statuts", "À venir · En vente · Vendu · Invendu · Invendu probable · Clôturé · Inconnu · Publié · À vérifier"],
    ["Type source", "Maison de vente · Agrégateur · Marchand · Marketplace · Publication"],
    ["Dédup", "Clé = URL (normalisée sans paramètres). Un même lot représenté (ex. Overgrown monument chez Fichter) est une nouvelle ligne avec une nouvelle date de vente."],
    ["Homonymes à exclure", "Jacques Bellange (graveur lorrain, 1575-1616) · Pierre-Antoine et Louis-Alexandre Bellangé (ébénistes) · Eugène Bellangé (fils, 1837-1895, à signaler séparément si intéressant) · Aurélien Bellanger (romancier)"],
    ["Source des données", "État des lieux du 25/08/2026 réalisé par Claude ; chaque ligne provient d'une page ouverte, URL en colonne M. Prix et estimations tels qu'affichés sur la page (frais compris ou non selon la mention)."],
]
for r in LEG:
    ws4.append(r)
ws4["A1"].font = Font(name="Arial", bold=True, size=13)
for row in ws4.iter_rows(min_row=3):
    row[0].font = Font(name="Arial", bold=True, size=10)
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cell.font.bold is not True:
            cell.font = Font(name="Arial", size=10)
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 110

import os; ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__))); wb.save(os.path.join(ROOT,"data","registre.xlsx"))
print(len(ITEMS), "entrées")
